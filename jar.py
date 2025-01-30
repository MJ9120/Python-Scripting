import openpyxl
import logging
import re

def separate_jar_files(file_path, output_file_path):
    try:
        # Configure logging
        logging.basicConfig(filename='separate_jar_files.log', level=logging.INFO, 
                            format='%(asctime)s - %(levelname)s - %(message)s')
        
        logging.info('Starting the process of separating .jar files...')
        
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        logging.info(f'Loaded workbook: {file_path}')
        
        # Regular expression to match .jar files (case insensitive)
        jar_regex = re.compile(r'([\w\-.]+\.jar)', re.IGNORECASE)
        
        # Iterate through all sheets
        for sheet in wb.worksheets:
            logging.info(f'Processing sheet: {sheet.title}')
            # Iterate through the rows and separate the .jar files
            for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
                files_value = sheet.cell(row=row, column=1).value  # Assuming 'Files' is in the first column
                if files_value:
                    # Find all matches for .jar files in the cell
                    jar_files = jar_regex.findall(files_value)
                    # Join multiple .jar files with a comma, or keep the first one
                    sheet.cell(row=row, column=2).value = ', '.join(jar_files) if jar_files else None
        
        # Save the updated Excel file
        wb.save(output_file_path)
        logging.info(f"The .jar files have been successfully separated and saved to {output_file_path}.")
        print(f"The .jar files have been successfully separated and saved to {output_file_path}.")
    
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")

# Define file paths
file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\Files.xlsx'
output_file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\Dell_separated.xlsx'

# Call the function to separate .jar files
separate_jar_files(file_path, output_file_path)
