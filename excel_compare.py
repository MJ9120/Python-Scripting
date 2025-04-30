import sys
from openpyxl import load_workbook

def get_header_index(headers, header_name):
    try:
        return headers.index(header_name.strip())
    except ValueError:
        print(f"Header '{header_name}' not found in the headers: {headers}")
        return None

def get_column_indices(headers, header_names):
    return {header_name: get_header_index(headers, header_name) for header_name in header_names}

if __name__ == "__main__":
    # Get file paths from arguments
    cc_old_path = sys.argv[1]
    cc_new_path = sys.argv[2]
    output_path = sys.argv[3]

    # Load the Excel workbooks
    cc_old_wb = load_workbook(cc_old_path)
    cc_new_wb = load_workbook(cc_new_path)

    # Process the 'Unapproved' sheet
    cc_old_ws_unapproved = cc_old_wb['Unapproved']
    cc_new_ws_unapproved = cc_new_wb['Unapproved']

    # Get header indices for the old and new 'Unapproved' sheets
    cc_old_headers_unapproved = [str(cell.value).strip() for cell in cc_old_ws_unapproved[1]]
    cc_new_headers_unapproved = [str(cell.value).strip() for cell in cc_new_ws_unapproved[1]]

    required_headers_unapproved = ['Name', 'Files', 'Match Type', 'Column B Status', 'Column C Status']
    cc_old_indices_unapproved = get_column_indices(cc_old_headers_unapproved, required_headers_unapproved)
    cc_new_indices_unapproved = get_column_indices(cc_new_headers_unapproved, required_headers_unapproved)

    # Check for missing headers
    if None in cc_old_indices_unapproved.values() or None in cc_new_indices_unapproved.values():
        print("Headers missing in 'Unapproved' sheets.")
        sys.exit(1)

    # Create dictionaries for 'Unapproved' data
    cc_old_statuses_unapproved = {}
    for row in cc_old_ws_unapproved.iter_rows(min_row=2, values_only=True):
        name = str(row[cc_old_indices_unapproved['Name']]).strip() if row[cc_old_indices_unapproved['Name']] else ''
        files = str(row[cc_old_indices_unapproved['Files']]).strip() if row[cc_old_indices_unapproved['Files']] else ''
        match_type = str(row[cc_old_indices_unapproved['Match Type']]).strip() if row[cc_old_indices_unapproved['Match Type']] else ''
        col_b_status = row[cc_old_indices_unapproved['Column B Status']]
        col_c_status = row[cc_old_indices_unapproved['Column C Status']]
        cc_old_statuses_unapproved[(name, files, match_type)] = (col_b_status, col_c_status)

    for row in cc_new_ws_unapproved.iter_rows(min_row=2):
        name = str(row[cc_new_indices_unapproved['Name']].value).strip() if row[cc_new_indices_unapproved['Name']].value else ''
        files = str(row[cc_new_indices_unapproved['Files']].value).strip() if row[cc_new_indices_unapproved['Files']].value else ''
        match_type = str(row[cc_new_indices_unapproved['Match Type']].value).strip() if row[cc_new_indices_unapproved['Match Type']].value else ''
        if (name, files, match_type) in cc_old_statuses_unapproved:
            col_b_status, col_c_status = cc_old_statuses_unapproved[(name, files, match_type)]
            row[cc_new_indices_unapproved['Column B Status']].value = col_b_status
            row[cc_new_indices_unapproved['Column C Status']].value = col_c_status
        else:
            row[cc_new_indices_unapproved['Column B Status']].value = 'Not Approved'
            row[cc_new_indices_unapproved['Column C Status']].value = 'Not Approved'

    # Save the updated workbook
    cc_new_wb.save(output_path)
    print(f"Updated report saved to {output_path}")
