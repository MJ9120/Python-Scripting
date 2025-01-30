import openpyxl
import logging

def separate_pkg_names(file_path, output_file_path):
    try:
        # Configure logging
        logging.basicConfig(filename='separate_pkg_names.log', level=logging.DEBUG, 
                            format='%(asctime)s - %(levelname)s - %(message)s')
        
        logging.info('Starting the process...')
        
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        logging.info(f'Loaded workbook: {file_path}')
        
        # List of file extensions to look for
        extensions = ['.deb', 'data.tar.zst', '.tar.gz', '.tar.xz', '.zip', '.tgz', 
                      '.tar.bz2', '.whl', '.jar', '.mod', '.dll']
        
        # Iterate through all sheets
        for sheet in wb.worksheets:
            logging.info(f'Processing sheet: {sheet.title}')
            # Iterate through the rows and separate the package names
            for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
                files_value = sheet.cell(row=row, column=1).value  # Assuming 'Files' is in the first column
                if files_value:
                    # Split by '!', '/', or '\\' to standardize delimiters
                    parts = files_value.replace('\\', '/').replace('!', '/').split('/')
                    
                    # Debug logging for parts
                    logging.debug(f'Row {row} parts: {parts}')
                    
                    # Filter parts that end with a valid extension and ignore paths with 'lockboxserver' (case-insensitive)
                    pkg_parts = [part for part in parts if any(part.lower().endswith(ext) for ext in extensions) 
                                 and 'lockboxserver' not in part.lower()]
                    
                    # Debug logging for filtered parts
                    logging.debug(f'Row {row} filtered parts: {pkg_parts}')
                    
                    # Get unique package names only
                    unique_pkg_parts = list(dict.fromkeys(pkg_parts))
                    
                    # Choose the unique package names separated by commas if multiple
                    pkg_name = ', '.join(unique_pkg_parts) if unique_pkg_parts else ''
                    
                    # Store the package name(s) in the second column
                    sheet.cell(row=row, column=2).value = pkg_name
        
        # Save the updated Excel file
        wb.save(output_file_path)
        logging.info(f"The package names have been successfully separated and saved to {output_file_path}.")
        print(f"The package names have been successfully separated and saved to {output_file_path}.")
    
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")

# Define file paths
file_path = r'C:\Users\Manjunath_MJalade\OneDrive - Dell Technologies\Desktop\Excel\New folder (2)\Files.xlsx'
output_file_path = r'C:\Users\Manjunath_MJalade\OneDrive - Dell Technologies\Desktop\Excel\New folder (2)\File_updated.xlsx'

# Call the function to separate package names
separate_pkg_names(file_path, output_file_path)