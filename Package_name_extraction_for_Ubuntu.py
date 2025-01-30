import openpyxl
import logging

def separate_pkg_names(file_path, output_file_path):
    try:
        # Configure logging
        logging.basicConfig(filename='separate_pkg_names.log', level=logging.INFO, 
                            format='%(asctime)s - %(levelname)s - %(message)s')
        
        logging.info('Starting the process...')
        
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        logging.info(f'Loaded workbook: {file_path}')
        
        # Iterate through all sheets
        for sheet in wb.worksheets:
            logging.info(f'Processing sheet: {sheet.title}')
            # Iterate through the rows and separate the package names
            for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
                files_value = sheet.cell(row=row, column=1).value  # Assuming 'Files' is in the first column
                if files_value:
                    # Extract the package name based on file extensions
                    parts = files_value.split('!')
                    pkg_name = None
                    for part in parts:
                        if any(ext in part for ext in ['.deb', 'data.tar.zst', '.tar.gz', '.tar.xz', '.zip', '.tgz', '.tar.bz2', '.whl']):
                            pkg_name = part.split('/')[-1]
                            if '.deb' in pkg_name:
                                break
                    if not pkg_name:
                        # Handle cases like ms-sdk and node path 0.12.7
                        parts = files_value.split('/')
                        for i in range(len(parts) - 1):
                            if any(ext in parts[i] for ext in ['.deb', 'data.tar.zst', '.tar.gz', '.tar.xz', '.zip', '.tgz', '.tar.bz2', '.whl']):
                                pkg_name = parts[i]
                                if '.deb' in pkg_name:
                                    break
                            elif '@' in parts[i] and len(parts) > i + 1 and parts[i + 1].replace('.', '').isdigit():
                                pkg_name = f"{parts[i]} {parts[i + 1]}"
                                break
                            elif '-' in parts[i] and len(parts) > i + 1 and parts[i + 1].replace('.', '').isdigit():
                                pkg_name = f"{parts[i]} {parts[i + 1]}"
                                break
                            elif 'node' in parts[i] and len(parts) > i + 1 and parts[i + 1].replace('.', '').isdigit():
                                pkg_name = f"node {parts[i + 1]}"
                                break
                            elif 'ms-sdk' in parts[i]:
                                pkg_name = parts[i]
                                break
                            elif 'code_coverage' in parts[i] and len(parts) > i + 1 and parts[i + 1].replace('.', '').isdigit():
                                pkg_name = f"{parts[i]} {parts[i + 1]}"
                                break
                            elif 'notification-system22' in parts[i] and len(parts) > i + 1 and '.dist-info' in parts[i + 1]:
                                pkg_name = parts[i + 1]
                                break
                            elif 'notification-system' in parts[i] and len(parts) > i + 1 and '.dist-info' in parts[i + 1]:
                                pkg_name = parts[i + 1]
                                break
                            elif 'mesa' in parts[i]:
                                pkg_name = 'mesa'
                                break
                            elif 'gnome-shell' in parts[i]:
                                pkg_name = 'gnome-shell'
                                break
                        if not pkg_name:
                            # Handle cases like notify2-0.3.1.dist-info specifically
                            for part in parts:
                                if '.dist-info' in part:
                                    pkg_name = part.split('/')[-1]
                                    break
                        if not pkg_name:
                            # If no specific package name is found, take the entire file path as package name
                            pkg_name = files_value
                    sheet.cell(row=row, column=2).value = pkg_name  # Assuming 'Pkg Name' should be in the second column
        
        # Save the updated Excel file
        wb.save(output_file_path)
        logging.info(f"The package names have been successfully separated and saved to {output_file_path}.")
        print(f"The package names have been successfully separated and saved to {output_file_path}.")
    
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")

# Define file paths
file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\Files.xlsx'
output_file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\DHC_Files.xlsx'

# Call the function to separate package names
separate_pkg_names(file_path, output_file_path)
