import openpyxl
import logging
import re

def separate_files(file_path, output_file_path):
    try:
        # Configure logging
        logging.basicConfig(filename='separate_files.log', level=logging.INFO, 
                            format='%(asctime)s - %(levelname)s - %(message)s')
        
        logging.info('Starting the process of separating files...')
        
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        logging.info(f'Loaded workbook: {file_path}')
        
        # Regular expression to match .jar files (case insensitive)
        jar_regex = re.compile(r'([\w\-.]+\.jar)', re.IGNORECASE)
        
        # Regular expression to match package/library paths (e.g., BouncyCastle.NetCoreSdk/1.9.7)
        package_regex = re.compile(r'([\w.-]+/[\d.]+)')
        
        # Regular expression to match specific file names with extensions (e.g., .dll, .exe, .cab)
        file_regex = re.compile(r'([\w\-.]+\.(dll|exe|cab))', re.IGNORECASE)
        
        # Iterate through all sheets
        for sheet in wb.worksheets:
            logging.info(f'Processing sheet: {sheet.title}')
            # Iterate through the rows and separate the files
            for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
                files_value = sheet.cell(row=row, column=1).value  # Assuming 'Files' is in the first column
                if files_value:
                    # Find all matches for .jar files in the cell
                    jar_files = jar_regex.findall(files_value)
                    # Find all matches for package paths in the cell
                    package_files = package_regex.findall(files_value)
                    # Find all matches for specific file names with extensions
                    file_matches = file_regex.findall(files_value)
                    extracted_files = jar_files + [match[0] for match in file_matches]
                    
                    # Filter out unwanted file paths, keeping only the necessary ones
                    filtered_files = []
                    for f in extracted_files:
                        if re.search(r'windowsdesktop-runtime-[\d.]+-win-x64\.exe\.cab$', f):
                            filtered_files.append(f)
                    
                    # Keep only relevant package paths, excluding UnifiedWyseDeviceAgent
                    relevant_packages = [pkg for pkg in package_files if 'UnifiedWyseDeviceAgent' not in pkg]
                    filtered_files.extend(relevant_packages)
                    
                    # Store the result in the second column
                    sheet.cell(row=row, column=2).value = ', '.join(filtered_files) if filtered_files else None
        
        # Save the updated Excel file
        wb.save(output_file_path)
        logging.info(f"The files have been successfully separated and saved to {output_file_path}.")
        print(f"The files have been successfully separated and saved to {output_file_path}.")
    
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")

# Define file paths
file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\Files.xlsx'
output_file_path = 'C:\\Users\\Manjunath_MJalade\\OneDrive - Dell Technologies\\Desktop\\Excel\\Dell_separated.xlsx'

# Call the function to separate files
separate_files(file_path, output_file_path)
