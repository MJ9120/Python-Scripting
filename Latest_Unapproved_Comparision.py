from openpyxl import load_workbook

# Load the Excel files
cc_old_path = r'C:\Users\Manjunath_MJalade\OneDrive - Dell Technologies\Desktop\Excel\CC_Old.xlsx'
cc_new_path = r'C:\Users\Manjunath_MJalade\OneDrive - Dell Technologies\Desktop\Excel\CC_New.xlsx'

cc_old_wb = load_workbook(cc_old_path)
cc_new_wb = load_workbook(cc_new_path)

cc_old_ws = cc_old_wb['Unapproved']
cc_new_ws = cc_new_wb['Unapproved']

# Print out the column names to verify
print("Columns in CC_Old.xlsx:")
for cell in cc_old_ws[1]:
    print(f"'{cell.value}'")

print("\nColumns in CC_New.xlsx:")
for cell in cc_new_ws[1]:
    print(f"'{cell.value}'")

# Get the column indices for the required columns, stripping any extra spaces
columns = {cell.value.strip(): cell.column for cell in cc_old_ws[1]}
name_col = columns['Name']
files_col = columns['Files']
match_type_col = columns['Match Type']
col_b_status_col = columns['Column B Status']
col_c_status_col = columns['Column C Status']

# Iterate through the rows in the new worksheet
for new_row in cc_new_ws.iter_rows(min_row=2, values_only=False):
    name = new_row[name_col - 1].value
    files = new_row[files_col - 1].value
    match_type = new_row[match_type_col - 1].value
    
    # Find matching rows in the old worksheet based on 'Name', 'Files', and 'Match Type' columns
    matching_row = None
    for old_row in cc_old_ws.iter_rows(min_row=2, values_only=False):
        if (old_row[name_col - 1].value == name and 
            old_row[files_col - 1].value == files and 
            old_row[match_type_col - 1].value == match_type):
            matching_row = old_row
            break
    
    # If a matching row is found, update 'Column B Status' and 'Column C Status' in the new worksheet
    if matching_row:
        new_row[col_b_status_col - 1].value = matching_row[col_b_status_col - 1].value
        new_row[col_c_status_col - 1].value = matching_row[col_c_status_col - 1].value
    else:
        # If no matching row is found, set 'Column B Status' and 'Column C Status' to 'Not Approved'
        new_row[col_b_status_col - 1].value = 'Not Approved'
        new_row[col_c_status_col - 1].value = 'Not Approved'

# Save the updated new workbook to a new Excel file
cc_new_wb.save(r'C:\Users\Manjunath_MJalade\OneDrive - Dell Technologies\Desktop\Excel\CC_New_Updated.xlsx')

print("The 'Column B Status' and 'Column C Status' have been updated in CC_New_Updated.xlsx where 'Name', 'Files', and 'Match Type' columns matched. If not matched, set to 'Not Approved'.")







